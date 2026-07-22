import openpyxl
import pandas as pd
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Optional


COLUMN_KEYWORDS = {
    "referencia": ["referencia"],
    "desc_item": ["desc. item", "desc item", "descripcion item", "item resumen"],
    "bodega": ["bodega"],
    "desc_ubicacion": ["desc. ubicacion", "desc ubicacion", "ubicacion"],
    "desc_bodega": ["desc. bodega", "desc bodega", "nombre bodega"],
    "lote": ["lote"],
    "um": ["u.m.", "u.m", "unidad medida", "unidad de medida"],
    "cant_disponible": ["cant. disponible", "cant disponible", "cantidad disponible"],
    "existencia": ["existencia"],
    "cant_comprometida": ["cant. comprometida", "cant comprometida", "cantidad comprometida", "cant. comprom", "cant comprom"],
    "precio_unitario": ["precio unitario", "precio"],
    "valor_total": ["valor total"],
    "margen": ["margen"],
    "linea": ["linea", "línea"],
    "sub_linea": ["sub-linea", "sub linea", "sublínea"],
    "diseno": ["diseño", "diseno", "dise?o"],
    "costo_prom_total": ["costo prom. total", "costo prom total"],
    "desc_co": ["desc. c.o. bodega", "desc c.o. bodega", "desc. c.o", "desc c.o", "desc. co"],
    "planeacion": ["planeacion"],
    "canal": ["canal"],
    "categoria": ["categoria"],
    "estado": ["estado"],
    "fecha_ultima_entrada": ["fecha ultima entrada", "fecha última entrada"],
    "fecha_ultima_salida": ["fecha ultima salida", "fecha última salida"],
    "cliente": ["cliente"],
    "proveedor": ["proveedor"],
    "detalle_ext1": ["detalle ext. 1", "detalle ext 1"],
    "detalle_ext2": ["detalle ext. 2", "detalle ext 2"],
    "costo_prom_unit_ins": ["costo prom. unit. (ins)", "costo prom unit ins"],
    "costo_prom_tot_ins": ["costo prom. tot. (ins)", "costo prom tot ins"],
    "ubicacion": ["ubicacion"],
}


def _normalize(text: str) -> str:
    result = text.strip().lower().replace(".", " ").replace("_", " ")
    while "  " in result:
        result = result.replace("  ", " ")
    return result


def _match_column(header: str, keywords: list[str]) -> bool:
    norm = _normalize(header)
    return any(_normalize(kw) in norm for kw in keywords)


def detect_columns(headers: list[str]) -> dict[str, Optional[int]]:
    detected = {}
    for key, keywords in COLUMN_KEYWORDS.items():
        detected[key] = None
        for idx, h in enumerate(headers):
            if h and _match_column(str(h), keywords):
                detected[key] = idx
                break
    return detected


def find_data_sheet(wb: openpyxl.Workbook) -> tuple[Optional[str], Optional[str]]:
    pivot_sheet = None
    data_sheet = None
    max_rows = 0

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row and ws.max_row > max_rows:
            max_rows = ws.max_row
            data_sheet = name

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row and ws.max_row < 50 and ws.max_column and ws.max_column <= 15:
            pivot_sheet = name
            break

    return data_sheet, pivot_sheet


def read_pivot_table(ws) -> dict:
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        rows.append(list(row))

    title = None
    filter_value = None
    headers_row = None
    data_rows = []

    for i, row in enumerate(rows):
        non_none = [c for c in row if c is not None]
        if len(non_none) >= 2 and isinstance(non_none[0], str) and "bodega" in _normalize(str(non_none[0])):
            title = str(non_none[0])
            filter_value = str(non_none[1]) if len(non_none) > 1 else None
            continue
        if any(c and isinstance(c, str) and "etiqueta" in _normalize(str(c)) for c in row if c):
            headers_row = row
            continue
        if any(c and isinstance(c, str) and "total" in _normalize(str(c)) for c in row if c):
            continue
        if headers_row is not None and any(c is not None for c in row):
            data_rows.append(row)

    return {
        "title": title,
        "filter_value": filter_value,
        "headers": headers_row,
        "data": data_rows,
    }


def _parse_pivot_from_df(df: pd.DataFrame) -> dict:
    title = None
    filter_value = None
    headers_row = None
    data_rows = []

    for i, row in df.iterrows():
        vals = [row[j] for j in range(len(df.columns))]
        non_none = [c for c in vals if pd.notna(c)]
        if len(non_none) >= 2 and isinstance(non_none[0], str) and "bodega" in _normalize(str(non_none[0])):
            title = str(non_none[0])
            filter_value = str(non_none[1]) if len(non_none) > 1 else None
            continue
        if any(pd.notna(c) and isinstance(c, str) and "etiqueta" in _normalize(str(c)) for c in vals):
            headers_row = [c if pd.notna(c) else None for c in vals]
            continue
        if any(pd.notna(c) and isinstance(c, str) and "total" in _normalize(str(c)) for c in vals):
            continue
        if headers_row is not None and any(pd.notna(c) for c in vals):
            data_rows.append([c if pd.notna(c) else None for c in vals])

    return {
        "title": title,
        "filter_value": filter_value,
        "headers": headers_row,
        "data": data_rows,
    }


def _safe_sum(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns:
        return 0.0
    return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())


def _safe_mean(df: pd.DataFrame, col: str) -> float:
    if col not in df.columns:
        return 0.0
    val = pd.to_numeric(df[col], errors="coerce").fillna(0).mean()
    return float(val) if pd.notna(val) else 0.0


def _safe_sum_from_series(series: pd.Series) -> float:
    return float(pd.to_numeric(series, errors="coerce").fillna(0).sum())


def compute_bodega_metrics(df: pd.DataFrame, col_map: dict[str, Optional[int]]) -> list[dict]:
    if "desc_bodega" not in df.columns and "bodega" not in df.columns:
        return []

    bodega_col = "desc_bodega" if "desc_bodega" in df.columns else "bodega"
    df_valid = df.dropna(subset=[bodega_col]).copy()

    str_cols = [c for c in ["desc_bodega", "linea", "sub_linea", "categoria", "canal", "estado", "proveedor", "desc_co", "referencia"] if c in df_valid.columns]
    for col in str_cols:
        df_valid[col] = df_valid[col].astype(str).str.strip()
        df_valid[col] = df_valid[col].replace({"": None, "nan": None, "None": None})

    num_cols = [c for c in ["existencia", "cant_comprometida", "cant_disponible", "valor_total", "costo_prom_total", "margen", "precio_unitario"] if c in df_valid.columns]
    for col in num_cols:
        df_valid[col] = pd.to_numeric(df_valid[col], errors="coerce").fillna(0)

    sub_dims = {
        "linea": ["existencia", "cant_comprometida", "cant_disponible", "valor_total", "margen"],
        "categoria": ["existencia", "cant_comprometida", "valor_total", "margen"],
        "canal": ["existencia", "cant_comprometida", "valor_total"],
        "estado": ["existencia", "cant_comprometida", "valor_total"],
        "proveedor": ["existencia", "cant_comprometida", "valor_total", "margen"],
    }

    sub_agg_results = {}
    for dim_name, agg_cols in sub_dims.items():
        if dim_name not in df_valid.columns:
            continue
        valid = df_valid.dropna(subset=[dim_name])
        if valid.empty:
            continue
        agg_dict = {}
        for c in agg_cols:
            agg_dict[c] = "mean" if c == "margen" else "sum"
        grouped = valid.groupby([bodega_col, dim_name], observed=False).agg(agg_dict).reset_index()
        grouped["registros"] = valid.groupby([bodega_col, dim_name], observed=False).size().values
        sub_agg_results[dim_name] = grouped

    if "desc_item" in df_valid.columns:
        valid_items = df_valid.dropna(subset=["desc_item"])
        if not valid_items.empty:
            items_grouped = valid_items.groupby([bodega_col, "desc_item"], observed=False).agg(
                existencia=("existencia", "sum"),
                comprometida=("cant_comprometida", "sum"),
                valor=("valor_total", "sum"),
                precio=("precio_unitario", "mean"),
                registros=("desc_item", "count"),
            ).reset_index()
            top_items_per_bodega = items_grouped.sort_values("valor", ascending=False).groupby(bodega_col).head(10)
            sub_agg_results["items"] = top_items_per_bodega

    bodega_agg = df_valid.groupby(bodega_col, observed=False).agg(
        existencia=("existencia", "sum"),
        cant_comprometida=("cant_comprometida", "sum"),
        cant_disponible=("cant_disponible", "sum"),
        valor_total=("valor_total", "sum"),
        costo_prom_total=("costo_prom_total", "sum"),
        margen_promedio=("margen", "mean"),
        precio_unitario=("precio_unitario", "mean"),
        total_registros=(bodega_col, "count"),
    ).reset_index()

    if "referencia" in df_valid.columns:
        ref_counts = df_valid.dropna(subset=["referencia"]).groupby(bodega_col, observed=False)["referencia"].nunique().reset_index()
        ref_counts.columns = [bodega_col, "refs_unicas"]
        bodega_agg = bodega_agg.merge(ref_counts, on=bodega_col, how="left")
        bodega_agg["refs_unicas"] = bodega_agg["refs_unicas"].fillna(0).astype(int)
    else:
        bodega_agg["refs_unicas"] = 0

    if "desc_item" in df_valid.columns:
        item_counts = df_valid.dropna(subset=["desc_item"]).groupby(bodega_col, observed=False)["desc_item"].nunique().reset_index()
        item_counts.columns = [bodega_col, "items_unicos"]
        bodega_agg = bodega_agg.merge(item_counts, on=bodega_col, how="left")
        bodega_agg["items_unicos"] = bodega_agg["items_unicos"].fillna(0).astype(int)
    else:
        bodega_agg["items_unicos"] = 0

    if "fecha_ultima_salida" in df_valid.columns:
        sin_mov = df_valid.groupby(bodega_col, observed=False)["fecha_ultima_salida"].apply(lambda x: x.isna().sum()).reset_index()
        sin_mov.columns = [bodega_col, "sin_movimiento"]
        bodega_agg = bodega_agg.merge(sin_mov, on=bodega_col, how="left")
        bodega_agg["sin_movimiento"] = bodega_agg["sin_movimiento"].fillna(0).astype(int)
    else:
        bodega_agg["sin_movimiento"] = 0

    if "bodega" in df_valid.columns and bodega_col != "bodega":
        codigo_map = df_valid.dropna(subset=["bodega"]).groupby(bodega_col, observed=False)["bodega"].first().reset_index()
        codigo_map.columns = [bodega_col, "bodega_codigo"]
        bodega_agg = bodega_agg.merge(codigo_map, on=bodega_col, how="left")
        bodega_agg["bodega_codigo"] = bodega_agg["bodega_codigo"].fillna("").astype(str)
    else:
        bodega_agg["bodega_codigo"] = bodega_agg[bodega_col]

    results = []
    for _, row in bodega_agg.iterrows():
        bodega_name = str(row[bodega_col]).strip()
        if not bodega_name or bodega_name == "None":
            continue
        existencia = row["existencia"]
        cant_comprometida = row["cant_comprometida"]
        compromiso_pct = (cant_comprometida / existencia * 100) if existencia > 0 else 0

        por_linea = {}
        if "linea" in sub_agg_results:
            sub = sub_agg_results["linea"][sub_agg_results["linea"][bodega_col] == bodega_name]
            for _, sr in sub.iterrows():
                k = str(sr["linea"]).strip()
                if k and k != "None":
                    por_linea[k] = {
                        "existencia": round(sr["existencia"], 0),
                        "comprometida": round(sr["cant_comprometida"], 0),
                        "disponible": round(sr.get("cant_disponible", sr["existencia"] - sr["cant_comprometida"]), 0),
                        "valor_total": round(sr["valor_total"], 0),
                        "margen": round(sr["margen"], 1),
                        "registros": int(sr["registros"]),
                    }

        por_categoria = {}
        if "categoria" in sub_agg_results:
            sub = sub_agg_results["categoria"][sub_agg_results["categoria"][bodega_col] == bodega_name]
            for _, sr in sub.iterrows():
                k = str(sr["categoria"]).strip()
                if k and k != "None":
                    por_categoria[k] = {
                        "existencia": round(sr["existencia"], 0),
                        "comprometida": round(sr["cant_comprometida"], 0),
                        "valor_total": round(sr["valor_total"], 0),
                        "margen": round(sr["margen"], 1),
                        "registros": int(sr["registros"]),
                    }

        por_canal = {}
        if "canal" in sub_agg_results:
            sub = sub_agg_results["canal"][sub_agg_results["canal"][bodega_col] == bodega_name]
            for _, sr in sub.iterrows():
                k = str(sr["canal"]).strip()
                if k and k != "None":
                    por_canal[k] = {
                        "existencia": round(sr["existencia"], 0),
                        "comprometida": round(sr["cant_comprometida"], 0),
                        "valor_total": round(sr["valor_total"], 0),
                        "registros": int(sr["registros"]),
                    }

        por_estado = {}
        if "estado" in sub_agg_results:
            sub = sub_agg_results["estado"][sub_agg_results["estado"][bodega_col] == bodega_name]
            for _, sr in sub.iterrows():
                k = str(sr["estado"]).strip()
                if k and k != "None":
                    por_estado[k] = {
                        "existencia": round(sr["existencia"], 0),
                        "comprometida": round(sr["cant_comprometida"], 0),
                        "valor_total": round(sr["valor_total"], 0),
                        "registros": int(sr["registros"]),
                    }

        por_proveedor = {}
        if "proveedor" in sub_agg_results:
            sub = sub_agg_results["proveedor"][sub_agg_results["proveedor"][bodega_col] == bodega_name]
            for _, sr in sub.iterrows():
                k = str(sr["proveedor"]).strip()
                if k and k != "None":
                    por_proveedor[k] = {
                        "existencia": round(sr["existencia"], 0),
                        "comprometida": round(sr["cant_comprometida"], 0),
                        "valor_total": round(sr["valor_total"], 0),
                        "margen": round(sr["margen"], 1),
                        "registros": int(sr["registros"]),
                    }

        top_items = []
        if "items" in sub_agg_results:
            sub = sub_agg_results["items"][sub_agg_results["items"][bodega_col] == bodega_name]
            for _, sr in sub.iterrows():
                top_items.append({
                    "item": str(sr["desc_item"]),
                    "existencia": round(sr["existencia"], 0),
                    "comprometida": round(sr["comprometida"], 0),
                    "valor_total": round(sr["valor"], 0),
                    "precio_unitario": round(sr["precio"], 0),
                    "registros": int(sr["registros"]),
                })

        results.append({
            "bodega": bodega_name,
            "bodega_codigo": str(row["bodega_codigo"]),
            "existencia": round(existencia, 0),
            "cant_comprometida": round(cant_comprometida, 0),
            "cant_disponible": round(row["cant_disponible"], 0),
            "valor_total": round(row["valor_total"], 0),
            "costo_prom_total": round(row["costo_prom_total"], 0),
            "margen_promedio": round(row["margen_promedio"], 1),
            "precio_unitario": round(row["precio_unitario"], 0),
            "compromiso_pct": round(compromiso_pct, 1),
            "total_registros": int(row["total_registros"]),
            "refs_unicas": int(row["refs_unicas"]),
            "items_unicos": int(row["items_unicos"]),
            "sin_movimiento": int(row["sin_movimiento"]),
            "por_linea": por_linea,
            "por_categoria": por_categoria,
            "por_canal": por_canal,
            "por_estado": por_estado,
            "por_proveedor": por_proveedor,
            "top_items": top_items,
        })

    results.sort(key=lambda x: x["valor_total"], reverse=True)
    return results


def _safe_mean_from_series(series: pd.Series) -> float:
    val = pd.to_numeric(series, errors="coerce").fillna(0).mean()
    return float(val) if pd.notna(val) else 0.0


def build_global_summary(metrics: list[dict]) -> dict:
    total_existencia = sum(m["existencia"] for m in metrics)
    total_comprometida = sum(m["cant_comprometida"] for m in metrics)
    total_disponible = sum(m["cant_disponible"] for m in metrics)
    total_valor = sum(m["valor_total"] for m in metrics)
    total_registros = sum(m["total_registros"] for m in metrics)
    total_refs = sum(m["refs_unicas"] for m in metrics)
    total_items = sum(m["items_unicos"] for m in metrics)
    total_sin_movimiento = sum(m["sin_movimiento"] for m in metrics)
    compromiso_global = (total_comprometida / total_existencia * 100) if total_existencia > 0 else 0

    margenes = [m["margen_promedio"] for m in metrics if m["margen_promedio"] > 0]
    margen_promedio = sum(margenes) / len(margenes) if margenes else 0

    precios = [m["precio_unitario"] for m in metrics if m["precio_unitario"] > 0]
    precio_promedio = sum(precios) / len(precios) if precios else 0

    linea_total = {}
    for m in metrics:
        for linea, data in m.get("por_linea", {}).items():
            if linea not in linea_total:
                linea_total[linea] = {"existencia": 0, "comprometida": 0, "valor_total": 0, "registros": 0}
            linea_total[linea]["existencia"] += data["existencia"]
            linea_total[linea]["comprometida"] += data["comprometida"]
            linea_total[linea]["valor_total"] += data["valor_total"]
            linea_total[linea]["registros"] += data["registros"]

    categoria_total = {}
    for m in metrics:
        for cat, data in m.get("por_categoria", {}).items():
            if cat not in categoria_total:
                categoria_total[cat] = {"existencia": 0, "comprometida": 0, "valor_total": 0, "registros": 0}
            categoria_total[cat]["existencia"] += data["existencia"]
            categoria_total[cat]["comprometida"] += data["comprometida"]
            categoria_total[cat]["valor_total"] += data["valor_total"]
            categoria_total[cat]["registros"] += data["registros"]

    canal_total = {}
    for m in metrics:
        for canal, data in m.get("por_canal", {}).items():
            if canal not in canal_total:
                canal_total[canal] = {"existencia": 0, "comprometida": 0, "valor_total": 0, "registros": 0}
            canal_total[canal]["existencia"] += data["existencia"]
            canal_total[canal]["comprometida"] += data["comprometida"]
            canal_total[canal]["valor_total"] += data["valor_total"]
            canal_total[canal]["registros"] += data["registros"]

    estado_total = {}
    for m in metrics:
        for estado, data in m.get("por_estado", {}).items():
            if estado not in estado_total:
                estado_total[estado] = {"existencia": 0, "comprometida": 0, "valor_total": 0, "registros": 0}
            estado_total[estado]["existencia"] += data["existencia"]
            estado_total[estado]["comprometida"] += data["comprometida"]
            estado_total[estado]["valor_total"] += data["valor_total"]
            estado_total[estado]["registros"] += data["registros"]

    return {
        "total_bodegas": len(metrics),
        "total_existencia": round(total_existencia, 0),
        "total_comprometida": round(total_comprometida, 0),
        "total_disponible": round(total_disponible, 0),
        "total_valor": round(total_valor, 0),
        "total_registros": total_registros,
        "total_refs_unicas": total_refs,
        "total_items_unicos": total_items,
        "total_sin_movimiento": total_sin_movimiento,
        "compromiso_global_pct": round(compromiso_global, 1),
        "margen_promedio": round(margen_promedio, 1),
        "precio_unitario_promedio": round(precio_promedio, 0),
        "linea_total": linea_total,
        "categoria_total": categoria_total,
        "canal_total": canal_total,
        "estado_total": estado_total,
    }


def build_bodega_prompt(bodega_data: dict, global_summary: dict, bodega_filter: str = "") -> str:
    bodega_info = f"\n- Filtro de bodega: {bodega_filter}" if bodega_filter else ""

    avg_valor = global_summary["total_valor"] / global_summary["total_bodegas"] if global_summary["total_bodegas"] > 0 else 0
    valor_ratio = (bodega_data["valor_total"] / global_summary["total_valor"] * 100) if global_summary["total_valor"] > 0 else 0

    linea_text = ""
    if bodega_data.get("por_linea"):
        linea_text = "\n".join(
            f"    - {l}: existencia {d['existencia']:,.0f}, comprometida {d['comprometida']:,.0f}, valor ${d['valor_total']:,.0f}, margen {d['margen']:.1f}%"
            for l, d in bodega_data["por_linea"].items()
        )
    else:
        linea_text = "    (Sin datos de lineas)"

    categoria_text = ""
    if bodega_data.get("por_categoria"):
        categoria_text = "\n".join(
            f"    - {c}: existencia {d['existencia']:,.0f}, comprometida {d['comprometida']:,.0f}, valor ${d['valor_total']:,.0f}, margen {d['margen']:.1f}%"
            for c, d in bodega_data["por_categoria"].items()
        )
    else:
        categoria_text = "    (Sin datos de categorias)"

    canal_text = ""
    if bodega_data.get("por_canal"):
        canal_text = "\n".join(
            f"    - {c}: existencia {d['existencia']:,.0f}, comprometida {d['comprometida']:,.0f}, valor ${d['valor_total']:,.0f}"
            for c, d in bodega_data["por_canal"].items()
        )
    else:
        canal_text = "    (Sin datos de canales)"

    estado_text = ""
    if bodega_data.get("por_estado"):
        estado_text = "\n".join(
            f"    - {e}: existencia {d['existencia']:,.0f}, comprometida {d['comprometida']:,.0f}, valor ${d['valor_total']:,.0f}"
            for e, d in bodega_data["por_estado"].items()
        )
    else:
        estado_text = "    (Sin datos de estados)"

    proveedor_text = ""
    if bodega_data.get("por_proveedor"):
        proveedor_text = "\n".join(
            f"    - {p}: existencia {d['existencia']:,.0f}, comprometida {d['comprometida']:,.0f}, valor ${d['valor_total']:,.0f}, margen {d['margen']:.1f}%"
            for p, d in bodega_data["por_proveedor"].items()
        )
    else:
        proveedor_text = "    (Sin datos de proveedores)"

    items_text = ""
    if bodega_data.get("top_items"):
        items_text = "\n".join(
            f"    - {i['item']}: existencia {i['existencia']:,.0f}, comprometida {i['comprometida']:,.0f}, valor ${i['valor_total']:,.0f}, precio ${i['precio_unitario']:,.0f}"
            for i in bodega_data["top_items"]
        )
    else:
        items_text = "    (Sin datos de items)"

    linea_team_text = ""
    if global_summary.get("linea_total"):
        sorted_lineas = sorted(global_summary["linea_total"].items(), key=lambda x: x[1]["valor_total"], reverse=True)[:10]
        linea_team_text = "\n".join(
            f"    - {l}: existencia {d['existencia']:,.0f}, valor ${d['valor_total']:,.0f}"
            for l, d in sorted_lineas
        )

    cat_team_text = ""
    if global_summary.get("categoria_total"):
        cat_team_text = "\n".join(
            f"    - {c}: existencia {d['existencia']:,.0f}, valor ${d['valor_total']:,.0f}"
            for c, d in global_summary["categoria_total"].items()
        )

    prompt = f"""Eres un analista experto en gestion de inventario y logistica para una empresa de muebles y construccion. Genera un informe detallado y profesional para la siguiente bodega.

## DATOS DE LA BODEGA
- Nombre: {bodega_data['bodega']}{bodega_info}

## METRICAS PRINCIPALES
- Existencia total: {bodega_data['existencia']:,.0f}
- Cantidad comprometida: {bodega_data['cant_comprometida']:,.0f}
- Cantidad disponible: {bodega_data['cant_disponible']:,.0f}
- % de Compromiso (comprometida/existencia): {bodega_data['compromiso_pct']}%
- Total de registros: {bodega_data['total_registros']}
- Referencias unicas: {bodega_data['refs_unicas']}
- Items unicos: {bodega_data['items_unicos']}
- Items sin movimiento: {bodega_data['sin_movimiento']}

## VALORES MONETARIOS Y MARGEN
- Valor total del inventario: ${bodega_data['valor_total']:,.0f}
- Costo promedio total: ${bodega_data['costo_prom_total']:,.0f}
- Precio unitario promedio: ${bodega_data['precio_unitario']:,.0f}
- Margen promedio: {bodega_data['margen_promedio']:.1f}%
- Porcentaje del valor total del sistema: {valor_ratio:.1f}%

## COMPARACION CON EL SISTEMA
- Promedio valor por bodega: ${avg_valor:,.0f}
- Total existencia del sistema: {global_summary['total_existencia']:,.0f}
- Total comprometida del sistema: {global_summary['total_comprometida']:,.0f}
- Compromiso global: {global_summary['compromiso_global_pct']}%
- Margen promedio del sistema: {global_summary['margen_promedio']:.1f}%

## LINEAS DE PRODUCTO EN ESTA BODEGA
{linea_text}

## CATEGORIAS EN ESTA BODEGA
{categoria_text}

## CANALES EN ESTA BODEGA
{canal_text}

## ESTADOS EN ESTA BODEGA
{estado_text}

## PROVEEDORES EN ESTA BODEGA
{proveedor_text}

## TOP ITEMS POR VALOR
{items_text}

## LINEAS DEL SISTEMA COMPLETO (Top 10 por valor)
{linea_team_text}

## CATEGORIAS DEL SISTEMA COMPLETO
{cat_team_text}

---

## INSTRUCCIONES
Genera un informe profesional usando formato Markdown. IMPORTANTE: Usa tablas Markdown para presentar datos numericos. Ejemplo:
| Metrica | Valor |
|---|---|
| Existencia | 5,000 |

### 1. RESUMEN EJECUTIVO
Tabla resumen con los datos principales de la bodega en formato | Metrica | Valor |.

### 2. ANALISIS DE INVENTARIO
Tabla comparativa: Bodega vs Promedio Sistema. Columnas: Metrica, Bodega, Promedio Sistema, Diferencia.

### 3. COMPROMISO Y DISPONIBILIDAD
Analisis del porcentaje de compromiso. Alertas si esta por encima del 70%.

### 4. VALORES MONETARIOS Y MARGEN
Tabla con valores: Categoria, Valor Bodega, Valor Sistema, % Participacion.

### 5. ANALISIS POR LINEAS Y CATEGORIAS
Tabla de las lineas principales con sus metricas. Analisis de categorias.

### 6. PROVEEDORES Y CANALES
Tabla de proveedores y canales con sus metricas.

### 7. ITEMS CLAVE Y ALERTAS
Top items por valor. Alertas de: items sin movimiento, compromiso alto, margen bajo.

### 8. RECOMENDACIONES ESPECIFICAS
Tabla con: #, Recomendacion, Accion, Prioridad, Impacto Esperado.

Sé directo, profesional y basado en datos numericos especificos."""

    return prompt


def process_excel(filepath: str, canal: str = None, categoria: str = None, estado: str = None, linea: str = None) -> dict:
    try:
        xls = pd.ExcelFile(filepath, engine='calamine')
    except Exception:
        try:
            xls = pd.ExcelFile(filepath, engine='openpyxl')
        except Exception as e:
            raise ValueError(f"No se pudo abrir el archivo Excel. Verifica que no este corrupto o abierto en otro programa: {e}")

    sheet_names = xls.sheet_names
    if not sheet_names:
        raise ValueError("El archivo Excel no tiene hojas.")

    data_sheet_name = None
    pivot_sheet_name = None
    max_rows = 0

    sheet_dfs = {}
    for name in sheet_names:
        try:
            df_check = pd.read_excel(xls, sheet_name=name, header=None)
            sheet_dfs[name] = df_check
            row_count = len(df_check)
            if row_count > max_rows:
                max_rows = row_count
                data_sheet_name = name
        except Exception:
            continue

    for name, df_check in sheet_dfs.items():
        if name != data_sheet_name and 0 < len(df_check) < 50 and len(df_check.columns) <= 15:
            pivot_sheet_name = name
            break

    pivot_data = None
    if pivot_sheet_name and pivot_sheet_name in sheet_dfs:
        pivot_data = _parse_pivot_from_df(sheet_dfs[pivot_sheet_name])

    if not data_sheet_name:
        raise ValueError("No se encontro una hoja con datos. Verifica que el archivo tenga al menos una hoja con datos.")

    if data_sheet_name in sheet_dfs:
        df_raw = sheet_dfs[data_sheet_name]
        if df_raw.shape[0] > 0:
            new_cols = [str(c).strip() if pd.notna(c) else '' for c in df_raw.iloc[0]]
            df_raw = df_raw.iloc[1:].reset_index(drop=True)
            df_raw.columns = new_cols
    else:
        df_raw = pd.read_excel(xls, sheet_name=data_sheet_name, header=0)

    headers = list(df_raw.columns)

    del sheet_dfs

    col_map = detect_columns(headers)

    rename_map = {}
    for key, idx in col_map.items():
        if idx is not None and idx < len(df_raw.columns):
            rename_map[df_raw.columns[idx]] = key
    df_raw = df_raw.rename(columns=rename_map)

    bodega_filter = ""
    if pivot_data and pivot_data.get("filter_value"):
        bodega_filter = pivot_data["filter_value"]

    df_filtered = df_raw

    if canal and "canal" in df_filtered.columns:
        df_filtered = df_filtered[df_filtered["canal"].astype(str).str.strip().str.lower() == canal.strip().lower()]
    if categoria and "categoria" in df_filtered.columns:
        df_filtered = df_filtered[df_filtered["categoria"].astype(str).str.strip().str.lower() == categoria.strip().lower()]
    if estado and "estado" in df_filtered.columns:
        df_filtered = df_filtered[df_filtered["estado"].astype(str).str.strip().str.lower() == estado.strip().lower()]
    if linea and "linea" in df_filtered.columns:
        df_filtered = df_filtered[df_filtered["linea"].astype(str).str.strip().str.lower() == linea.strip().lower()]

    bodega_metrics = compute_bodega_metrics(df_filtered, col_map)
    global_summary = build_global_summary(bodega_metrics)

    return {
        "pivot_table": pivot_data,
        "col_map": {k: v for k, v in col_map.items() if v is not None},
        "bodega_metrics": bodega_metrics,
        "global_summary": global_summary,
        "bodega_filter": bodega_filter,
        "total_raw_rows": len(df_raw),
        "total_filtered_rows": len(df_filtered),
        "detected_columns": {k: (headers[v] if v is not None else None) for k, v in col_map.items()},
    }


DISPLAY_COLS = ["bodega", "desc_bodega", "referencia", "desc_item", "linea",
                "sub_linea", "categoria", "canal", "estado", "proveedor",
                "existencia", "cant_comprometida", "cant_disponible",
                "valor_total", "precio_unitario", "lote", "ubicacion"]


def search_inventory(query: str, df: pd.DataFrame, limit: int = 50) -> list[dict]:
    if not query or not query.strip():
        return []
    q = query.strip().lower()
    str_cols = [c for c in df.columns if df[c].dtype == object or c in DISPLAY_COLS]
    mask = pd.Series(False, index=df.index)
    for col in str_cols:
        if col in df.columns:
            mask |= df[col].astype(str).str.lower().str.contains(q, na=False)
    matched = df[mask].head(limit)
    results = []
    for _, row in matched.iterrows():
        item = {}
        for col in DISPLAY_COLS:
            if col in row.index:
                val = row[col]
                if pd.isna(val):
                    item[col] = None
                elif isinstance(val, (int, float)):
                    item[col] = round(float(val), 2)
                else:
                    item[col] = str(val).strip()
        results.append(item)
    return results
